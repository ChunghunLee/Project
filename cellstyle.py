from openpyxl.styles import colors,Color,Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
                
wb = Workbook()
ws = wb.active
my_cell = ws['B1']
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
fill = PatternFill("solid", fgColor="DDDDDD")
font = Font(name='HY헤드라인M',size = 17 , bold=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")
