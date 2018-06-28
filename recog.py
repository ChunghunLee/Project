import cv2
import numpy as np
import pytesseract
import datetime
import os

from PIL import Image, ImageEnhance

from openpyxl import *
from openpyxl.reader.excel import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors,Color,Font,Alignment,Side,Border,PatternFill
from cellstyle import style_range  # 셀 스타일 함수 사용하기위함
from requests import get        #서버 응답 모듈
from db2 import DBconnect       #DB 연결 모듈

class Recognition:
    
    def download(self, url,file_name):      #서버에서 불러온 사진(url)을 /python에 file_name으로  저장
    # Write Binary 모드로 연다,                   
        with open(file_name, "wb") as file:
            response = get(url)
            file.write(response.content)
            
            
    def ExtractNumber(self):
        #받아온 이미지(orig) 크기 변환 함수
        orig = Image.open('test.jpg') #'C:/test/test.jpg'
        width,height = 650,487  #사진마다 최적의 해상도가 있음
        
        if orig.size [0] < orig.size[1]:
            width,height = height,width

            
        cn = orig.convert('RGB')
        contrast = ImageEnhance.Brightness(cn)
        contrast = contrast.enhance(1.4)  #밝기 조절(어두우면 높이기)
     
        resized_img = contrast.resize((width, height), Image.ANTIALIAS)

        
        resized_img.save('res_img.jpg')
        
        Number = 'res_img.jpg'
        img=cv2.imread(Number,cv2.IMREAD_COLOR)
        copy_img=img.copy()
        img2=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        cv2.imwrite('gray.jpg',img2)
        #컬러 이미지를 GRAY로 바꿈
        blur = cv2.GaussianBlur(img2,(3,3),0)
        cv2.imwrite('blur.jpg',blur)
        #윤곽선 잘잡게 가우시안필터 적용
        canny=cv2.Canny(blur,100,200)
        cv2.imwrite('canny.jpg',canny)
        #canny 엣지 검출
        cnts,contours,hierarchy  = cv2.findContours(canny, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
 
        box1=[]
        f_count=0
        select=0
        plate_width=0       

        for i in range(len(contours)):
               cnt=contours[i]          
               area = cv2.contourArea(cnt)
               x,y,w,h = cv2.boundingRect(cnt)
               rect_area=w*h  #인식할 네모의 넓이 (높이 * 너비)
               aspect_ratio = float(w)/h # ratio = width/height
                  
               if  (aspect_ratio>=0.2)and(aspect_ratio<=1.2)and(rect_area>=100)and(rect_area<=700): # 700 기본값 
                    cv2.rectangle(img,(x,y),(x+w+1,y+h+1),(0,255,0),1)
                    box1.append(cv2.boundingRect(cnt))

        for i in range(len(box1)): ##버블 정렬(번호판 크기의 정사각형 인식을 위해)
               for j in range(len(box1)-(i+1)):
                    if box1[j][0]>box1[j+1][0]:
                         temp=box1[j]
                         box1[j]=box1[j+1]
                         box1[j+1]=temp

        #그 정사각형들 사이에서 길이를 측정하는 for문
        for m in range(len(box1)):
               count=0
               for n in range(m+1,(len(box1)-1)):
                    delta_x=abs(box1[n+1][0]-box1[m][0])
                    if delta_x > 150:
                         break
                    delta_y =abs(box1[n+1][1]-box1[m][1])
                    if delta_x ==0:
                         delta_x=1
                    if delta_y ==0:
                         delta_y=1           
                    gradient =float(delta_y) /float(delta_x)
                    if gradient<0.25:
                        count=count+1
               #number_plate의 크기를 측정한다.     
               if count > f_count:
                    select = m
                    f_count = count;
                    plate_width=delta_x
        cv2.imwrite('snake.jpg',img)


        number_plate=copy_img[box1[select][1]-10:box1[select][3]+box1[select][1]+5, box1[select][0]-10:160+box1[select][0]]
        #[x:y , z:w] 형식, x는 위->아래로 얼마나 자를지, y는 밑->위로 얼마나 자를지, z,w는 각각 좌->우/우->좌 (+일수록 덜 자름)
        resize_plate=cv2.resize(number_plate,None,fx=1.8,fy=1.8,interpolation=cv2.INTER_CUBIC+cv2.INTER_LINEAR)
        plate_gray=cv2.cvtColor(resize_plate,cv2.COLOR_BGR2GRAY)
        ret,th_plate = cv2.threshold(plate_gray,150,255,cv2.THRESH_BINARY)
        #임계값 150을 기준으로, 작으면 흑 크면 백(255)로 적용,BINARY는 흑/백  
        cv2.imwrite('plate_th.jpg',th_plate)
        kernel = np.ones((3,3),np.uint8)
        er_plate = cv2.erode(th_plate,kernel,iterations=1)
        er_invplate = er_plate
        cv2.imwrite('er_plate.jpg',er_invplate)
        
        cv2.imshow('image',img)
        cv2.imshow('image2',number_plate)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        global result
        result = pytesseract.image_to_string(Image.open('er_plate.jpg'), lang='kor')
        tmp_result = result.replace(" ","") 
        temp = 'c:/test/recog_img/'+tmp_result+'.jpg'
        
        print(temp)
        resized_img.save(temp)
        return(result.replace(" ",""))

        
    def timecal(self):              # 엑셀에 시간함수
        dt = datetime.datetime.now()
        time = dt.strftime('%Y-%m-%d %H:%M:%S')
        return(time)
    
url="http://ec2-13-209-68-158.ap-northeast-2.compute.amazonaws.com/drone_img/test.jpg"      #서버에서 불러올 이미지 이름
file_name = 'test.jpg'       #컴퓨터 내에 저장할 이름
    
recogtest=Recognition()
recogtest.download(url,file_name)           # 서버에서 파일을 다운로드
result = recogtest.ExtractNumber()          # result에 해당 파일 번호판 인식 결과 저장

time = recogtest.timecal()                  #시간 표시
                  
DBconnect(result,'Woo','cbnuroot123','CBNU')    # DB 연결

#print(time)
print(result)

cmd = 'java -jar c:/test/test.jar'          # GUI 호출
print (cmd)
os.system(cmd)

#결과 값들을 Excel로 만들기

wb = Workbook()  #load_workbook('carnum.xlsx') 로 불러오기
ws = wb.active
ws.title = 'Car_num'

cal = ws['C2']
cal2 = ws['E2']
multi_cells1 = ws['B2':'B5']
multi_cells2 = ws['G2':'G5']

ws['B1'] = '차량번호'
ws['E1'] = '적발 시간'

thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
fill = PatternFill("solid", fgColor="DDDDDD")
font = Font(name='HY헤드라인M',size = 17 , bold=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")
fill1 = PatternFill('solid', fgColor='FFFF66')

style_range(ws,'B1:D1',border=border,fill=fill,font=font,alignment=al)
style_range(ws,'E1:G1',border=border,fill=fill,font=font,alignment=al)
style_range(ws,'E2:G2',border=border,fill=None,font=None,alignment=al)
style_range(ws,'E3:G3',border=border,fill=None,font=None,alignment=al)
style_range(ws,'E4:G4',border=border,fill=None,font=None,alignment=al)
style_range(ws,'E5:G5',border=border,fill=None,font=None,alignment=al)
style_range(ws,'B2:D2',border=border,fill=None,font=None,alignment=al)
style_range(ws,'B3:D3',border=border,fill=None,font=None,alignment=al)
style_range(ws,'B4:D4',border=border,fill=None,font=None,alignment=al)
style_range(ws,'B5:D5',border=border,fill=None,font=None,alignment=al)
style_range(ws,'A1:A14',border=border,fill=fill1,font=None,alignment=None)


ws['B2'] = result
ws['E2'] = time

'''
data = []
for row in ws.rows:
    data.append([row[0].value, row[1].value , row[2].value])
'''
wb.save('carnum.xlsx')


